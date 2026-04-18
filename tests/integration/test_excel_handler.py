import pytest
import pandas as pd
from openpyxl import Workbook, load_workbook
import os
from src.app.infra.excel_handler import ExcelHandler

@pytest.fixture
def temp_excel_files(tmp_path):
    # Create a dummy origin data file
    data_dir = tmp_path / "data"
    input_dir = data_dir / "input"
    output_dir = data_dir / "output"
    input_dir.mkdir(parents=True)
    output_dir.mkdir(parents=True)
    
    origin_path = input_dir / "medicao.xlsx"
    df = pd.DataFrame({
        'Data': [pd.Timestamp('2026-03-01'), pd.Timestamp('2026-03-02')],
        'Descrição do serviço': ['Serviço 1', 'Serviço 2'],
        'Bairro': ['Centro', 'Bairro 2']
    })
    
    with pd.ExcelWriter(origin_path) as writer:
        df.to_excel(writer, sheet_name='manual', index=False)
        df.to_excel(writer, sheet_name='semaforica', index=False)

    # Create a dummy template file
    template_path = input_dir / "template.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws["A1"] = "Data:"
    ws["B10"] = "Semaforica:"
    ws["B15"] = "Manual:"
    wb.save(template_path)
    
    return {
        "origin": str(origin_path),
        "template": str(template_path),
        "output_dir": str(output_dir)
    }

def test_gerar_diario_completo(temp_excel_files, monkeypatch):
    config = {
        'projeto': {'ano': 2026, 'mes': 3},
        'arquivos': {
            'dados_origem': temp_excel_files['origin'],
            'template_ativo': temp_excel_files['template']
        },
        'posicoes': {'celula_data': 'E3'},
        'colunas': {
            'data': 'Data',
            'servico': 'Descrição do serviço',
            'bairro': 'Bairro'
        },
        'mapeamento': {
            'semaforica': 'B10',
            'manual': 'B15'
        }
    }
    
    # Patch the output directory to use our temp one
    # Since ExcelHandler uses a hardcoded 'data/output/' string
    # We might need to mock or change the working directory
    
    original_cwd = os.getcwd()
    os.chdir(os.path.dirname(temp_excel_files['output_dir'])) # Change to parent of 'output'
    
    try:
        # Create 'data/output' inside the temp dir if it doesn't exist relative to CWD
        os.makedirs("data/output", exist_ok=True)
        
        handler = ExcelHandler(config)
        output_path = handler.gerar_diario_completo()
        
        assert os.path.exists(output_path)
        
        # Verify generated content
        wb_result = load_workbook(output_path)
        
        # Should have 31 sheets (March 2026)
        assert len(wb_result.sheetnames) == 31
        assert "01-03" in wb_result.sheetnames
        assert "31-03" in wb_result.sheetnames
        
        ws_01 = wb_result["01-03"]
        assert ws_01["E3"].value == "01/03/2026"
        assert ws_01["B15"].value == "Serviço 1"
        assert ws_01["B10"].value == "Serviço 1"
        
    finally:
        os.chdir(original_cwd)

def test_gerar_diario_template_missing(tmp_path):
    # Create a dummy data file to pass the first read_excel check
    data_path = tmp_path / "data.xlsx"
    pd.DataFrame({'A': [1]}).to_excel(data_path)
    
    config = {
        'arquivos': {
            'dados_origem': str(data_path),
            'template_ativo': str(tmp_path / "non_existent.xlsx")
        }
    }
    handler = ExcelHandler(config)
    with pytest.raises(FileNotFoundError, match="O arquivo de template está vazio ou não existe"):
        handler.gerar_diario_completo()
