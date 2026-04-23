import pytest
import pandas as pd
from openpyxl import Workbook, load_workbook
from pathlib import Path
from unittest.mock import patch
from src.app.core.report_builder import ReportBuilder
from src.app.core.config_models import ReportConfig
from src.app.core import constants

@pytest.fixture
def temp_setup(tmp_path):
    # Setup directories
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    output_dir.mkdir()
    
    # Create origin data
    origin_path = input_dir / "medicao.xlsx"
    df = pd.DataFrame({
        'Data': [pd.Timestamp('2026-03-01'), pd.Timestamp('2026-03-02')],
        'Descrição do serviço': ['Serviço 1', 'Serviço 2'],
        'Bairro': ['Centro', 'Bairro 2']
    })
    
    with pd.ExcelWriter(origin_path) as writer:
        df.to_excel(writer, sheet_name='manual', index=False)
        df.to_excel(writer, sheet_name='semaforica', index=False)

    # Create template
    template_path = input_dir / "template.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws["A1"] = "Data:"
    ws["B10"] = "Semaforica:"
    ws["B15"] = "Manual:"
    wb.save(template_path)
    
    config = ReportConfig()
    config.projeto.ano = 2026
    config.projeto.mes = 3
    config.arquivos.dados_origem = str(origin_path)
    config.arquivos.template_ativo = str(template_path)
    config.posicoes.celula_data_atual = 'E3'
    config.extracao.colunas = ['Data', 'Bairro', 'Descrição do serviço']
    config.extracao.formato_final = '{Descrição do serviço}'
    config.mapeamento = {
        'semaforica': 'B10',
        'manual': 'B15'
    }

    return config, output_dir

def test_report_builder_integration(temp_setup):
    config, output_dir = temp_setup
    output_path = output_dir / "resultado.xlsx"
    
    builder = ReportBuilder(config)
    builder.build(output_path)
    
    assert output_path.exists()
    
    wb_result = load_workbook(output_path)
    # March has 31 days
    assert len(wb_result.sheetnames) == 31
    assert "01-03" in wb_result.sheetnames
    
    ws_01 = wb_result["01-03"]
    assert ws_01["E3"].value == "01/03/2026"
    assert ws_01["B15"].value == "Serviço 1"

def test_deduplicacao_servicos(temp_setup):
    config, output_dir = temp_setup
    origin_path = Path(config.arquivos.dados_origem)
    output_path = output_dir / "resultado_dup.xlsx"

    # Data with duplicates
    df_duplicado = pd.DataFrame({
        'Data': [pd.Timestamp('2026-03-01'), pd.Timestamp('2026-03-01')],
        'Descrição do serviço': ['PINTURA DE MEIO FIO', 'PINTURA DE MEIO FIO'],
        'Bairro': ['Centro', 'Centro']
    })
    with pd.ExcelWriter(origin_path) as writer:
        df_duplicado.to_excel(writer, sheet_name='manual', index=False)

    builder = ReportBuilder(config)
    builder.build(output_path)
    
    wb_result = load_workbook(output_path)
    ws_01 = wb_result["01-03"]
    # Should be deduplicated and title-cased
    assert ws_01["B15"].value == "Pintura de Meio Fio"

@pytest.mark.parametrize("ano, mes, esperado", [
    (2024, 2, 29),  # Leap year
    (2025, 2, 28),
    (2026, 1, 31),
])
def test_limite_dias_mes(temp_setup, ano, mes, esperado):
    config, output_dir = temp_setup
    config.projeto.ano = ano
    config.projeto.mes = mes
    output_path = output_dir / f"teste_{ano}_{mes}.xlsx"

    builder = ReportBuilder(config)
    builder.build(output_path)
    
    wb = load_workbook(output_path)
    assert len(wb.sheetnames) == esperado
